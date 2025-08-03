from flask import render_template, request, jsonify, redirect, flash, url_for, send_file
from flask_login import login_required, current_user
from app.reports import reports
from app.models import User, Company, Project, TimeEntry
from app import db, csrf
from sqlalchemy import func, and_, text
from datetime import datetime, timedelta
from app import csrf
import openpyxl
import zipfile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Register fonts for Serbian characters
def register_serbian_fonts():
    """Register fonts that support Serbian characters"""
    try:
        # Try to register DejaVu fonts which support Serbian characters
        pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf'))
        return 'DejaVuSans'
    except:
        try:
            # Fallback to Arial Unicode MS if available
            pdfmetrics.registerFont(TTFont('ArialUnicodeMS', 'ARIALUNI.TTF'))
            return 'ArialUnicodeMS'
        except:
            # Use built-in fonts that support Unicode
            return 'Helvetica'

# New date formatting helper functions (duplicated from main, should be refactored to a common utility)
def format_date_for_display(date_obj):
    """Convert date to DD.MM.YYYY format for display"""
    if isinstance(date_obj, str):
        date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
    return date_obj.strftime('%d.%m.%Y')

def format_date_for_input(date_obj):
    """Convert date to DD.MM.YYYY format for input fields"""
    if isinstance(date_obj, str):
        date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
    return date_obj.strftime('%d.%m.%Y')

def parse_date_from_input(date_str):
    """Parse date from DD.MM.YYYY format to date object"""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%d.%m.%Y').date()
    except ValueError:
        # Fallback to YYYY-MM-DD format for backward compatibility
        return datetime.strptime(date_str, '%Y-%m-%d').date()

def format_date_for_api(date_obj):
    """Convert date to DD.MM.YYYY format for API responses"""
    if isinstance(date_obj, str):
        date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
    return date_obj.strftime('%d.%m.%Y')

@reports.route('/')
@login_required
def index():
    # For regular users, show their personal report directly
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        # Get user's time entries with project and company information
        sql_query = """
            SELECT
                te.date,
                te.hours,
                te.description,
                p.name as project_name,
                p.id as project_id,
                c.name as company_name,
                c.id as company_id
            FROM time_entries te
            JOIN projects p ON te.project_id = p.id
            JOIN companies c ON p.company_id = c.id
            WHERE te.user_id = :user_id
        """
        params = {'user_id': current_user.id}

        if start_date:
            sql_query += " AND te.date >= :start_date"
            params['start_date'] = parse_date_from_input(start_date)

        if end_date:
            sql_query += " AND te.date <= :end_date"
            params['end_date'] = parse_date_from_input(end_date)

        sql_query += " ORDER BY c.name, p.name, te.date"

        result = db.session.execute(text(sql_query), params)
        entries = result.fetchall()

        # Group data by company and project
        report_data = {}
        total_hours = 0

        for entry in entries:
            company_name = entry.company_name
            project_name = entry.project_name
            project_id = entry.project_id
            company_id = entry.company_id

            if company_name not in report_data:
                report_data[company_name] = {
                    'company_id': company_id,
                    'total_hours': 0,
                    'projects': {}
                }

            if project_name not in report_data[company_name]['projects']:
                report_data[company_name]['projects'][project_name] = {
                    'project_id': project_id,
                    'total_hours': 0,
                    'daily_entries': {}
                }

            # Add daily entry
            date_str = entry.date.strftime('%d.%m.%Y')
            if date_str not in report_data[company_name]['projects'][project_name]['daily_entries']:
                report_data[company_name]['projects'][project_name]['daily_entries'][date_str] = {
                    'hours': 0,
                    'description': entry.description or ''
                }

            report_data[company_name]['projects'][project_name]['daily_entries'][date_str]['hours'] += float(entry.hours)
            report_data[company_name]['projects'][project_name]['total_hours'] += float(entry.hours)
            report_data[company_name]['total_hours'] += float(entry.hours)

            total_hours += float(entry.hours)

        # Create flat list of all entries for the table
        all_entries = []
        for entry in entries:
            all_entries.append({
                'date': entry.date,
                'company_name': entry.company_name,
                'project_name': entry.project_name,
                'hours': float(entry.hours),
                'description': entry.description
            })

        return render_template('reports/index.html',
                             report_data=report_data,
                             all_entries=all_entries,
                             total_hours=total_hours,
                             start_date=start_date,
                             end_date=end_date)

    # For admin users, show the regular reports index
    return render_template('reports/index.html')

@reports.route('/my-report')
@login_required
def my_report():
    """Personal report for regular users showing their data grouped by companies and projects"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get user's time entries with project and company information
    sql_query = """
        SELECT 
            te.date,
            te.hours,
            te.description,
            p.name as project_name,
            p.id as project_id,
            c.name as company_name,
            c.id as company_id
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        WHERE te.user_id = :user_id
    """
    params = {'user_id': current_user.id}
    
    if start_date:
        sql_query += " AND te.date >= :start_date"
        params['start_date'] = parse_date_from_input(start_date)
    
    if end_date:
        sql_query += " AND te.date <= :end_date"
        params['end_date'] = parse_date_from_input(end_date)
    
    sql_query += " ORDER BY c.name, p.name, te.date"
    
    result = db.session.execute(text(sql_query), params)
    entries = result.fetchall()
    
    # Group data by company and project
    report_data = {}
    total_hours = 0
    
    for entry in entries:
        company_name = entry.company_name
        project_name = entry.project_name
        project_id = entry.project_id
        company_id = entry.company_id
        
        if company_name not in report_data:
            report_data[company_name] = {
                'company_id': company_id,
                'total_hours': 0,
                'projects': {}
            }
        
        if project_name not in report_data[company_name]['projects']:
            report_data[company_name]['projects'][project_name] = {
                'project_id': project_id,
                'total_hours': 0,
                'daily_entries': {}
            }
        
        # Add daily entry
        date_str = entry.date.strftime('%d.%m.%Y')
        if date_str not in report_data[company_name]['projects'][project_name]['daily_entries']:
            report_data[company_name]['projects'][project_name]['daily_entries'][date_str] = {
                'hours': 0,
                'description': entry.description or ''
            }
        
        report_data[company_name]['projects'][project_name]['daily_entries'][date_str]['hours'] += float(entry.hours)
        report_data[company_name]['projects'][project_name]['total_hours'] += float(entry.hours)
        report_data[company_name]['total_hours'] += float(entry.hours)
        
        total_hours += float(entry.hours)
    
    # Create flat list of all entries for the table
    all_entries = []
    for entry in entries:
        all_entries.append({
            'date': entry.date,
            'company_name': entry.company_name,
            'project_name': entry.project_name,
            'hours': float(entry.hours),
            'description': entry.description
        })

    return render_template('reports/my_report.html', 
                         report_data=report_data,
                         all_entries=all_entries,
                         total_hours=total_hours,
                         start_date=start_date,
                         end_date=end_date)

@reports.route('/export/my-report/excel')
@login_required
@csrf.exempt
def export_my_report_excel():
    """Export personal report to Excel with grouped data"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Get user's time entries with project and company information
    sql_query = """
        SELECT
            te.date,
            te.hours,
            te.description,
            p.name as project_name,
            p.id as project_id,
            c.name as company_name,
            c.id as company_id
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        WHERE te.user_id = :user_id
    """
    params = {'user_id': current_user.id}

    if start_date:
        sql_query += " AND te.date >= :start_date"
        params['start_date'] = parse_date_from_input(start_date)

    if end_date:
        sql_query += " AND te.date <= :end_date"
        params['end_date'] = parse_date_from_input(end_date)

    sql_query += " ORDER BY c.name, p.name, te.date"

    result = db.session.execute(text(sql_query), params)
    entries = result.fetchall()

    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from decimal import Decimal

    wb = Workbook()
    ws = wb.active
    ws.title = "Moj izveštaj"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    company_header_font = Font(bold=True, color="FFFFFF")
    company_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    project_header_font = Font(bold=True, color="FFFFFF")
    project_header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = f"IZVEŠTAJ ZA KORISNIKA: {current_user.first_name} {current_user.last_name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    # Period
    period_text = "Period: "
    if start_date and end_date:
        period_text += f"{start_date} - {end_date}"
    elif start_date:
        period_text += f"od {start_date}"
    elif end_date:
        period_text += f"do {end_date}"
    else:
        period_text += "svi podaci"

    ws['A2'] = period_text
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')

    current_row = 4

    # Group data by company and project
    report_data = {}
    total_hours = 0

    for entry in entries:
        company_name = entry.company_name
        project_name = entry.project_name
        project_id = entry.project_id
        company_id = entry.company_id

        if company_name not in report_data:
            report_data[company_name] = {
                'company_id': company_id,
                'total_hours': 0,
                'projects': {}
            }

        if project_name not in report_data[company_name]['projects']:
            report_data[company_name]['projects'][project_name] = {
                'project_id': project_id,
                'total_hours': 0,
                'daily_entries': {}
            }

        # Add daily entry
        date_str = entry.date.strftime('%d.%m.%Y')
        if date_str not in report_data[company_name]['projects'][project_name]['daily_entries']:
            report_data[company_name]['projects'][project_name]['daily_entries'][date_str] = {
                'hours': 0,
                'description': entry.description or ''
            }

        report_data[company_name]['projects'][project_name]['daily_entries'][date_str]['hours'] += float(entry.hours)
        report_data[company_name]['projects'][project_name]['total_hours'] += float(entry.hours)
        report_data[company_name]['total_hours'] += float(entry.hours)

        total_hours += float(entry.hours)

    # Get all unique dates for the period
    all_dates = set()
    for company_name, company_data in report_data.items():
        for project_name, project_data in company_data['projects'].items():
            for date_str in project_data['daily_entries'].keys():
                all_dates.add(date_str)
    
    # Sort dates
    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
    
    # Write data to Excel with daily columns
    for company_name, company_data in report_data.items():
        # Company header
        ws[f'A{current_row}'] = f"KOMPANIJA: {company_name}"
        ws[f'A{current_row}'].font = company_header_font
        ws[f'A{current_row}'].fill = company_header_fill
        ws[f'A{current_row}'].border = border
        ws.merge_cells(f'A{current_row}:{get_column_letter(3 + len(sorted_dates))}{current_row}')
        current_row += 1

        for project_name, project_data in company_data['projects'].items():
            # Project header
            ws[f'A{current_row}'] = f"  PROJEKAT: {project_name}"
            ws[f'A{current_row}'].font = project_header_font
            ws[f'A{current_row}'].fill = project_header_fill
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'A{current_row}:{get_column_letter(3 + len(sorted_dates))}{current_row}')
            current_row += 1

            # Header row with dates
            ws[f'A{current_row}'] = "Kompanija"
            ws[f'B{current_row}'] = "Projekat"
            ws[f'C{current_row}'] = "Korisnik"
            
            # Add date columns
            for i, date_str in enumerate(sorted_dates):
                col_letter = get_column_letter(4 + i)
                ws[f'{col_letter}{current_row}'] = date_str
                ws[f'{col_letter}{current_row}'].font = header_font
                ws[f'{col_letter}{current_row}'].fill = header_fill
                ws[f'{col_letter}{current_row}'].border = border
            
            # Add total column
            total_col = get_column_letter(4 + len(sorted_dates))
            ws[f'{total_col}{current_row}'] = "UKUPNO"
            ws[f'{total_col}{current_row}'].font = header_font
            ws[f'{total_col}{current_row}'].fill = header_fill
            ws[f'{total_col}{current_row}'].border = border
            
            # Style header row
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = header_font
                ws[f'{col}{current_row}'].fill = header_fill
                ws[f'{col}{current_row}'].border = border
            current_row += 1

            # Data row
            ws[f'A{current_row}'] = company_name
            ws[f'B{current_row}'] = project_name
            ws[f'C{current_row}'] = f"{current_user.first_name} {current_user.last_name}"
            
            # Add hours for each date
            for i, date_str in enumerate(sorted_dates):
                col_letter = get_column_letter(4 + i)
                hours = project_data['daily_entries'].get(date_str, {}).get('hours', 0)
                ws[f'{col_letter}{current_row}'] = hours if hours > 0 else ""
                ws[f'{col_letter}{current_row}'].border = border
            
            # Add project total
            total_col = get_column_letter(4 + len(sorted_dates))
            ws[f'{total_col}{current_row}'] = project_data['total_hours']
            ws[f'{total_col}{current_row}'].font = Font(bold=True)
            ws[f'{total_col}{current_row}'].border = border
            
            # Style data row
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].border = border
            current_row += 2

        # Company total
        current_row += 1
        ws[f'A{current_row}'] = f"UKUPNO ZA KOMPANIJU: {company_name}"
        ws[f'A{current_row}'].font = total_font
        ws[f'A{current_row}'].fill = total_fill
        ws[f'A{current_row}'].border = border
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1

        # Company total row
        ws[f'A{current_row}'] = "UKUPNO"
        ws[f'B{current_row}'] = company_data['total_hours']
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].font = total_font
            ws[f'{col}{current_row}'].fill = total_fill
            ws[f'{col}{current_row}'].border = border
        current_row += 3  # Extra space between companies

    # Final summary table
    current_row += 2
    ws[f'A{current_row}'] = "UKUPAN PREGLED PO KOMPANIJAMA"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 2

    # Summary table header
    ws[f'A{current_row}'] = "Kompanija"
    ws[f'B{current_row}'] = "Naziv korisnika"
    ws[f'C{current_row}'] = "Broj radnih sati"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].font = header_font
        ws[f'{col}{current_row}'].fill = header_fill
        ws[f'{col}{current_row}'].border = border
    current_row += 1

    # Summary table data
    for company_name, company_data in report_data.items():
        ws[f'A{current_row}'] = company_name
        ws[f'B{current_row}'] = f"{current_user.first_name} {current_user.last_name}"
        ws[f'C{current_row}'] = company_data['total_hours']
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = border
        current_row += 1

    # Final total row
    ws[f'A{current_row}'] = "UKUPAN ZBIR ZA PERIOD"
    ws[f'B{current_row}'] = ""
    ws[f'C{current_row}'] = total_hours
    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].font = total_font
        ws[f'{col}{current_row}'].fill = total_fill
        ws[f'{col}{current_row}'].border = border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"moj_izvestaj_{current_user.username}"
    if start_date:
        filename += f"_od_{start_date.replace('.', '-')}"
    if end_date:
        filename += f"_do_{end_date.replace('.', '-')}"
    filename += ".xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports.route('/user-summary')
@login_required
def user_summary():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = TimeEntry.query
    
    if current_user.is_super_admin() or current_user.is_company_admin():
        # Admin sees all entries
        pass
    else:
        # Regular user sees only their entries
        query = query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    # Group by user only (unique users)
    results = db.session.execute(text("""
        SELECT 
            u.id as user_id,
            u.first_name, u.last_name, u.username,
            u.email,
            COUNT(DISTINCT p.id) as project_count,
            COUNT(DISTINCT c.id) as company_count,
            SUM(te.hours) as total_hours,
            COUNT(te.id) as entry_count
        FROM time_entries te
        JOIN users u ON te.user_id = u.id
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        WHERE 1=1
        """ + (" AND te.user_id = :user_id" if not (current_user.is_super_admin() or current_user.is_company_admin()) else "") + """
        """ + (" AND te.date >= :start_date" if start_date else "") + """
        """ + (" AND te.date <= :end_date" if end_date else "") + """
        GROUP BY u.id, u.first_name, u.last_name, u.username, u.email
        ORDER BY u.last_name, u.first_name
    """), {
        'user_id': current_user.id if not (current_user.is_super_admin() or current_user.is_company_admin()) else None,
        'start_date': parse_date_from_input(start_date) if start_date else None,
        'end_date': parse_date_from_input(end_date) if end_date else None
    })
    
    summary_data = []
    for row in results.fetchall():
        summary_data.append({
            'user_id': row.user_id,
            'user_name': f"{row.first_name} {row.last_name}",
            'username': row.username,
            'email': row.email,
            'project_count': row.project_count,
            'company_count': row.company_count,
            'total_hours': float(row.total_hours),
            'entry_count': row.entry_count
        })
    
    return render_template('reports/user_summary.html', summary_data=summary_data)

@reports.route('/project-summary')
@login_required
def project_summary():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = TimeEntry.query
    
    if current_user.is_super_admin() or current_user.is_company_admin():
        # Admin sees all entries
        pass
    else:
        # Regular user sees only their entries
        query = query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    # Group by project only (unique projects)
    results = db.session.execute(text("""
        SELECT 
            p.id as project_id,
            p.name as project_name,
            c.id as company_id,
            c.name as company_name,
            COUNT(DISTINCT u.id) as unique_users,
            SUM(te.hours) as total_hours,
            COUNT(te.id) as entry_count
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        JOIN users u ON te.user_id = u.id
        WHERE 1=1
        """ + (" AND te.user_id = :user_id" if not (current_user.is_super_admin() or current_user.is_company_admin()) else "") + """
        """ + (" AND te.date >= :start_date" if start_date else "") + """
        """ + (" AND te.date <= :end_date" if end_date else "") + """
        GROUP BY p.id, c.id
        ORDER BY c.name, p.name
    """), {
        'user_id': current_user.id if not (current_user.is_super_admin() or current_user.is_company_admin()) else None,
        'start_date': parse_date_from_input(start_date) if start_date else None,
        'end_date': parse_date_from_input(end_date) if end_date else None
    })
    
    summary_data = []
    for row in results.fetchall():
        summary_data.append({
            'project_id': row.project_id,
            'project_name': row.project_name,
            'company_id': row.company_id,
            'company_name': row.company_name,
            'unique_users': row.unique_users,
            'total_hours': float(row.total_hours),
            'entry_count': row.entry_count
        })
    
    return render_template('reports/project_summary.html', summary_data=summary_data)

@reports.route('/company-summary')
@login_required
def company_summary():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = TimeEntry.query
    
    if current_user.is_super_admin() or current_user.is_company_admin():
        # Admin sees all entries
        pass
    else:
        # Regular user sees only their entries
        query = query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    # Group by company only (unique companies)
    results = db.session.execute(text("""
        SELECT 
            c.id as company_id,
            c.name as company_name,
            COUNT(DISTINCT p.id) as project_count,
            COUNT(DISTINCT u.id) as user_count,
            SUM(te.hours) as total_hours,
            COUNT(te.id) as entry_count
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        JOIN users u ON te.user_id = u.id
        WHERE 1=1
        """ + (" AND te.user_id = :user_id" if not (current_user.is_super_admin() or current_user.is_company_admin()) else "") + """
        """ + (" AND te.date >= :start_date" if start_date else "") + """
        """ + (" AND te.date <= :end_date" if end_date else "") + """
        GROUP BY c.id, c.name
        ORDER BY c.name
    """), {
        'user_id': current_user.id if not (current_user.is_super_admin() or current_user.is_company_admin()) else None,
        'start_date': parse_date_from_input(start_date) if start_date else None,
        'end_date': parse_date_from_input(end_date) if end_date else None
    })
    
    summary_data = []
    for row in results.fetchall():
        summary_data.append({
            'company_id': row.company_id,
            'company_name': row.company_name,
            'project_count': row.project_count,
            'user_count': row.user_count,
            'total_hours': float(row.total_hours),
            'entry_count': row.entry_count
        })
    
    return render_template('reports/company_summary.html', summary_data=summary_data)

@reports.route('/daily-report')
@login_required
def daily_report():
    date_str = request.args.get('date', datetime.now().strftime('%d.%m.%Y'))
    date = parse_date_from_input(date_str)
    today = datetime.now().date()
    
    if current_user.is_super_admin() or current_user.is_company_admin():
        entries = TimeEntry.query.filter_by(date=date).all()
    else:
        entries = TimeEntry.query.filter_by(date=date, user_id=current_user.id).all()
    
    total_hours = sum(entry.hours for entry in entries)
    
    # Calculate additional statistics
    unique_users = len(set(entry.user_id for entry in entries))
    unique_projects = len(set(entry.project_id for entry in entries))
    
    # Calculate project summary
    project_summary = db.session.execute(text("""
        SELECT 
            p.id as project_id,
            p.name as project_name,
            c.name as company_name,
            SUM(te.hours) as total_hours,
            COUNT(DISTINCT te.user_id) as user_count
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        WHERE te.date = :date
        GROUP BY p.id, p.name, c.name
        ORDER BY total_hours DESC
    """), {'date': date}).fetchall()
    
    # Calculate user summary
    user_summary = db.session.execute(text("""
        SELECT 
            u.id as user_id,
            u.first_name, u.last_name, u.username,
            SUM(te.hours) as total_hours,
            COUNT(DISTINCT te.project_id) as project_count
        FROM time_entries te
        JOIN users u ON te.user_id = u.id
        WHERE te.date = :date
        GROUP BY u.id, u.first_name, u.last_name, u.username
        ORDER BY total_hours DESC
    """), {'date': date}).fetchall()
    
    return render_template('reports/daily_report.html', 
                         entries=entries, 
                         date=date, 
                         today=today,
                         total_hours=total_hours,
                         unique_users=unique_users,
                         unique_projects=unique_projects,
                         project_summary=project_summary,
                         user_summary=user_summary,
                         timedelta=timedelta)

@reports.route('/export/daily-report/<date_str>')
@login_required
@csrf.exempt
def export_daily_excel(date_str):
    """Export daily report to Excel"""
    date = parse_date_from_input(date_str)
    
    if current_user.is_super_admin() or current_user.is_company_admin():
        entries = TimeEntry.query.filter_by(date=date).all()
    else:
        entries = TimeEntry.query.filter_by(date=date, user_id=current_user.id).all()
    
    # Create Excel file
    wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
    ws = wb.create_sheet("Dnevni izveštaj")
    
    # Format worksheet
    start_row = format_worksheet(ws, f"Dnevni izveštaj - {format_date_for_display(date)}")
    
    # Add period information
    period_text = f"Datum: {format_date_for_display(date)}"
    ws.cell(row=4, column=1, value=period_text).font = Font(bold=True, italic=True)
    
    # Create headers
    headers = ['Korisnik', 'Projekat', 'Kompanija', 'Sati', 'Opis', 'Vreme unosa']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data
    current_row = 7
    for entry in entries:
        ws.cell(row=current_row, column=1, value=entry.user.get_full_name()).border = border
        ws.cell(row=current_row, column=2, value=entry.project.name).border = border
        ws.cell(row=current_row, column=3, value=entry.project.company.name).border = border
        ws.cell(row=current_row, column=4, value=entry.hours).border = border
        ws.cell(row=current_row, column=4).number_format = '0.00'
        ws.cell(row=current_row, column=5, value=entry.description or '').border = border
        ws.cell(row=current_row, column=6, value=entry.created_at.strftime('%H:%M')).border = border
        current_row += 1
    
    # Add summary
    current_row += 1
    total_hours = sum(entry.hours for entry in entries)
    ws.cell(row=current_row, column=1, value="UKUPNO").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=total_hours).font = Font(bold=True)
    ws.cell(row=current_row, column=4).number_format = '0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with date
    filename = f"dnevni_izvestaj_{date.strftime('%Y%m%d')}.xlsx"
    
    return send_file(excel_file, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@reports.route('/export/daily-report/<date_str>/pdf')
@login_required
@csrf.exempt
def export_daily_pdf(date_str):
    """Export daily report to PDF"""
    date = parse_date_from_input(date_str)
    
    if current_user.is_super_admin() or current_user.is_company_admin():
        entries = TimeEntry.query.filter_by(date=date).all()
    else:
        entries = TimeEntry.query.filter_by(date=date, user_id=current_user.id).all()
    
    # Create PDF
    pdf_file = BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
    
    # Register fonts for Serbian characters
    serbian_font = register_serbian_fonts()
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=serbian_font
    )
    
    # Build PDF content
    story = []
    
    # Title
    story.append(Paragraph(f"Dnevni izveštaj - {format_date_for_display(date)}", title_style))
    story.append(Spacer(1, 20))
    
    # Summary statistics
    total_hours = sum(entry.hours for entry in entries)
    unique_users = len(set(entry.user_id for entry in entries))
    unique_projects = len(set(entry.project_id for entry in entries))
    
    summary_data = [
        ['Ukupno sati', f"{total_hours:.2f}"],
        ['Korisnika', str(unique_users)],
        ['Projekata', str(unique_projects)],
        ['Unosa', str(len(entries))]
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), serbian_font),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Time entries table
    if entries:
        # Prepare table data
        table_data = [['Korisnik', 'Projekat', 'Kompanija', 'Sati', 'Opis']]
        
        for entry in entries:
            description = entry.description[:50] + '...' if entry.description and len(entry.description) > 50 else (entry.description or '')
            table_data.append([
                entry.user.get_full_name(),
                entry.project.name,
                entry.project.company.name,
                f"{entry.hours:.2f}",
                description
            ])
        
        # Add total row
        table_data.append(['UKUPNO', '', '', f"{total_hours:.2f}", ''])
        
        # Create table
        table = Table(table_data, colWidths=[120, 120, 100, 60, 200])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), serbian_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), f'{serbian_font}-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    
    # Generate filename with date
    filename = f"dnevni_izvestaj_{date.strftime('%Y%m%d')}.pdf"
    
    return send_file(pdf_file, as_attachment=True, download_name=filename, mimetype='application/pdf')

# Excel Export Functions
def create_excel_workbook():
    """Create a new Excel workbook with proper styling"""
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return wb, header_font, header_fill, header_alignment, border

def format_worksheet(ws, title):
    """Format worksheet with title and styling"""
    # Set title
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:Z1')
    
    # Add timestamp
    ws['A2'] = f"Generisano: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:Z2')
    
    return 4  # Return starting row for data

@reports.route('/export/project/<int:project_id>/pdf')
@login_required
@csrf.exempt
def export_project_pdf(project_id):
    """Export detailed project report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get project details
    project = Project.query.get_or_404(project_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin() or 
            current_user.is_project_admin(project_id)):
        flash('Nemate dozvolu za pristup ovom projektu', 'error')
        return redirect(url_for('reports.index'))
    
    # Get time entries for the project
    query = TimeEntry.query.filter_by(project_id=project_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(TimeEntry.date, TimeEntry.user_id).all()
    
    # Create PDF with landscape orientation for better width
    pdf_file = BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
    
    # Register fonts for Serbian characters
    serbian_font = register_serbian_fonts()
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=serbian_font
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        spaceBefore=20,
        fontName=serbian_font
    )
    
    # Build PDF content
    story = []
    
    # Title (centered)
    story.append(Paragraph(f"Izveštaj projekta: {project.name}", title_style))
    story.append(Spacer(1, 10))

    # Period information (centered)
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    story.append(Paragraph(period_text, title_style))
    story.append(Spacer(1, 20))

    # Project info table
    project_info = [
        ['Kompanija:', project.company.name],
        ['Opis:', project.description or 'Nema opisa'],
        ['Status:', project.status],
        ['Datum početka:', format_date_for_display(project.start_date)],
        ['Datum završetka:', format_date_for_display(project.end_date) if project.end_date else 'Nije definisan'],
        ['Budžet:', f"{float(project.budget):,.2f} RSD" if project.budget else 'Nije definisan']
    ]
    info_table = Table(project_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), serbian_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 30))

    # Page break for summary
    story.append(PageBreak())

    # Group entries by user for summary
    users = {}
    total_hours = 0
    for entry in entries:
        if entry.user_id not in users:
            users[entry.user_id] = {
                'name': entry.user.get_full_name(),
                'total_hours': 0
            }
        users[entry.user_id]['total_hours'] += entry.hours
        total_hours += entry.hours

    # Summary table po korisnicima
    summary_data = [['Korisnik', 'Ukupno sati']]
    for user in users.values():
        summary_data.append([user['name'], f"{user['total_hours']:.2f}"])
    summary_data.append(['UKUPNO', f"{total_hours:.2f}"])
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), serbian_font),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), serbian_font),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(Paragraph("Sumarna statistika po korisnicima", title_style))
    story.append(Spacer(1, 20))
    story.append(summary_table)

    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    # Create filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"projekat_{project.name.replace(' ', '_')}{date_suffix}.pdf"
    return send_file(
        pdf_file,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/project/<int:project_id>')
@login_required
@csrf.exempt
def export_project_excel(project_id):
    """Export detailed project report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get project details
    project = Project.query.get_or_404(project_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin() or 
            current_user.is_project_admin(project_id)):
        flash('Nemate dozvolu za pristup ovom projektu', 'error')
        return redirect(url_for('reports.index'))
    
    # Get time entries for the project
    query = TimeEntry.query.filter_by(project_id=project_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(TimeEntry.date, TimeEntry.user_id).all()
    
    # Create Excel workbook
    wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
    ws = wb.active
    ws.title = f"Projekat - {project.name[:20]}"
    
    # Format worksheet
    start_row = format_worksheet(ws, f"Izveštaj projekta: {project.name}")
    
    # Add period information
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
    
    # Get unique dates and users
    dates = sorted(list(set(entry.date for entry in entries)))
    users = {}
    for entry in entries:
        if entry.user_id not in users:
            users[entry.user_id] = entry.user.get_full_name()
    
    # Create headers (dates as rows, users as columns)
    headers = ['Datum'] + [user_name for user_name in users.values()] + ['UKUPNO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data (dates as rows, users as columns)
    current_row = start_row + 1
    user_totals = {user_id: 0 for user_id in users.keys()}
    
    for date in dates:
        ws.cell(row=current_row, column=1, value=format_date_for_display(date)).border = border
        date_total = 0
        
        for col, (user_id, user_name) in enumerate(users.items(), 2):
            # Find entry for this user and date
            entry = next((e for e in entries if e.user_id == user_id and e.date == date), None)
            hours = entry.hours if entry else 0
            user_totals[user_id] += hours
            date_total += hours
            
            cell = ws.cell(row=current_row, column=col, value=hours)
            cell.border = border
            if hours > 0:
                cell.number_format = '0.00'
        
        # Add date total
        total_cell = ws.cell(row=current_row, column=len(users) + 2, value=date_total)
        total_cell.border = border
        total_cell.font = Font(bold=True)
        total_cell.number_format = '0.00'
        
        current_row += 1
    
    # Add grand total row
    ws.cell(row=current_row, column=1, value="UKUPNO").font = Font(bold=True)
    ws.cell(row=current_row, column=1).border = border
    
    grand_total = 0
    for col, (user_id, user_name) in enumerate(users.items(), 2):
        user_total = user_totals[user_id]
        grand_total += user_total
        
        cell = ws.cell(row=current_row, column=col, value=user_total)
        cell.font = Font(bold=True)
        cell.border = border
        cell.number_format = '0.00'
    
    # Grand total
    grand_total_cell = ws.cell(row=current_row, column=len(users) + 2, value=grand_total)
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.border = border
    grand_total_cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"projekat_{project.name.replace(' ', '_')}{date_suffix}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/all-projects/zip')
@login_required
@csrf.exempt
def export_all_projects_zip():
    """Export all projects to ZIP with individual Excel and PDF files"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get all projects
    projects = Project.query.all()
    
    # Create ZIP file
    zip_file = BytesIO()
    with zipfile.ZipFile(zip_file, 'w') as zipf:
        for project in projects:
            # Get time entries for this project
            query = db.session.query(TimeEntry).filter(TimeEntry.project_id == project.id)
            if start_date:
                query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
            if end_date:
                query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
            
            entries = query.order_by(TimeEntry.date, TimeEntry.user_id).all()
            
            if not entries:
                continue  # Skip projects with no entries
            
            # Create Excel file for this project
            wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
            ws = wb.active
            ws.title = f"Projekat - {project.name[:20]}"
            
            # Format worksheet
            start_row = format_worksheet(ws, f"Izveštaj projekta: {project.name}")
            
            # Add period information
            period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
            ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
            
            # Get unique dates and users
            dates = sorted(list(set(entry.date for entry in entries)))
            users = {}
            for entry in entries:
                if entry.user_id not in users:
                    users[entry.user_id] = entry.user.get_full_name()
            
            # Create headers (dates as rows, users as columns)
            headers = ['Datum'] + [user_name for user_name in users.values()] + ['UKUPNO']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Fill data
            current_row = start_row + 1
            user_totals = {user_id: 0 for user_id in users.keys()}
            
            for date in dates:
                ws.cell(row=current_row, column=1, value=format_date_for_display(date)).border = border
                date_total = 0
                
                for col, (user_id, user_name) in enumerate(users.items(), 2):
                    entry = next((e for e in entries if e.user_id == user_id and e.date == date), None)
                    hours = entry.hours if entry else 0
                    user_totals[user_id] += hours
                    date_total += hours
                    
                    cell = ws.cell(row=current_row, column=col, value=hours)
                    cell.border = border
                    if hours > 0:
                        cell.number_format = '0.00'
                
                total_cell = ws.cell(row=current_row, column=len(users) + 2, value=date_total)
                total_cell.border = border
                total_cell.font = Font(bold=True)
                total_cell.number_format = '0.00'
                
                current_row += 1
            
            # Add grand total row
            ws.cell(row=current_row, column=1, value="UKUPNO").font = Font(bold=True)
            ws.cell(row=current_row, column=1).border = border
            
            grand_total = 0
            for col, (user_id, user_name) in enumerate(users.items(), 2):
                user_total = user_totals[user_id]
                grand_total += user_total
                
                cell = ws.cell(row=current_row, column=col, value=user_total)
                cell.font = Font(bold=True)
                cell.border = border
                cell.number_format = '0.00'
            
            grand_total_cell = ws.cell(row=current_row, column=len(users) + 2, value=grand_total)
            grand_total_cell.font = Font(bold=True)
            grand_total_cell.border = border
            grand_total_cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Create PDF for this project
            pdf_file = BytesIO()
            doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
            
            # Register fonts for Serbian characters
            serbian_font = register_serbian_fonts()
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName=serbian_font
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                spaceBefore=20,
                fontName=serbian_font
            )
            
            # Build PDF content
            story = []
            
            # Title (centered)
            story.append(Paragraph(f"Izveštaj projekta: {project.name}", title_style))
            story.append(Spacer(1, 10))
            
            # Period information (centered)
            period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
            story.append(Paragraph(period_text, title_style))
            story.append(Spacer(1, 20))
            
            # Project info table
            project_info = [
                ['Kompanija:', project.company.name],
                ['Opis:', project.description or 'Nema opisa'],
                ['Status:', project.status],
                ['Datum početka:', format_date_for_display(project.start_date)],
                ['Datum završetka:', format_date_for_display(project.end_date) if project.end_date else 'Nije definisan'],
                ['Budžet:', f"{float(project.budget):,.2f} RSD" if project.budget else 'Nije definisan']
            ]
            info_table = Table(project_info, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.grey),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), serbian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(info_table)
            story.append(Spacer(1, 30))
            
            # Add page break for summary
            story.append(PageBreak())
            
            # Group entries by user for summary
            users = {}
            total_hours = 0
            for entry in entries:
                if entry.user_id not in users:
                    users[entry.user_id] = {
                        'name': entry.user.get_full_name(),
                        'total_hours': 0
                    }
                users[entry.user_id]['total_hours'] += entry.hours
                total_hours += entry.hours
            
            # Summary table po korisnicima
            summary_data = [['Korisnik', 'Ukupno sati']]
            for user in users.values():
                summary_data.append([user['name'], f"{user['total_hours']:.2f}"])
            summary_data.append(['UKUPNO', f"{total_hours:.2f}"])
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), serbian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
                ('FONTNAME', (0, -1), (-1, -1), serbian_font),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(Paragraph("Sumarna statistika po korisnicima", title_style))
            story.append(Spacer(1, 20))
            story.append(summary_table)
            
            # Build PDF
            doc.build(story)
            pdf_file.seek(0)
            
            # Add files to ZIP with date range in filename
            date_suffix = ""
            if start_date or end_date:
                date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
            
            safe_project_name = project.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            excel_filename = f"projekat_{safe_project_name}{date_suffix}.xlsx"
            pdf_filename = f"projekat_{safe_project_name}{date_suffix}.pdf"
            
            zipf.writestr(excel_filename, excel_file.getvalue())
            zipf.writestr(pdf_filename, pdf_file.getvalue())
    
    zip_file.seek(0)
    
    # Create ZIP filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"svi_projekti{date_suffix}.zip"
    
    return send_file(
        zip_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/company/<int:company_id>')
@login_required
@csrf.exempt
def export_company_excel(company_id):
    """Export detailed company report to Excel with same format as all companies"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get company details
    company = Company.query.get_or_404(company_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get time entries for the company's projects
    query = db.session.query(TimeEntry).join(Project).filter(Project.company_id == company_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(Project.name, TimeEntry.date, TimeEntry.user_id).all()
    
    # Create Excel workbook
    wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
    ws = wb.active
    ws.title = f"Kompanija - {company.name[:20]}"
    
    # Format worksheet
    start_row = format_worksheet(ws, f"Izveštaj kompanije: {company.name}")
    
    # Add period information
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
    
    # Get unique dates and projects
    dates = sorted(list(set(entry.date for entry in entries)))
    projects = {}
    for entry in entries:
        if entry.project_id not in projects:
            projects[entry.project_id] = entry.project.name
    
    # Create headers (same format as all companies)
    headers = ['Projekat', 'Korisnik'] + [format_date_for_display(date) for date in dates] + ['Ukupno', 'Zarada']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data (same format as all companies)
    current_row = start_row + 1
    project_totals = {}
    user_earnings = {}  # Track earnings per user
    total_earnings = 0  # Track total earnings
    
    for project_id, project_name in projects.items():
        project_totals[project_id] = 0
        
        # Get users for this project
        project_users = {}
        for entry in entries:
            if entry.project_id == project_id and entry.user_id not in project_users:
                project_users[entry.user_id] = entry.user.get_full_name()
        
        for user_id, user_name in project_users.items():
            # Get user's hourly rate
            user = User.query.get(user_id)
            hourly_rate = float(user.hourly_rate) if user and user.hourly_rate else 0.0
            
            ws.cell(row=current_row, column=1, value=project_name).border = border
            ws.cell(row=current_row, column=2, value=user_name).border = border
            
            user_project_hours = 0
            for col, date in enumerate(dates, 3):
                # Find entry for this user, project and date
                entry = next((e for e in entries if e.user_id == user_id and e.project_id == project_id and e.date == date), None)
                hours = entry.hours if entry else 0
                user_project_hours += hours
                project_totals[project_id] += hours
                
                cell = ws.cell(row=current_row, column=col, value=round(hours) if hours > 0 else 0)
                cell.border = border
                if hours > 0:
                    cell.number_format = '0'
            
            # Add user total for this project
            total_cell = ws.cell(row=current_row, column=len(dates) + 3, value=round(user_project_hours))
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.number_format = '0'
            
            # Calculate and add user earnings for this project
            user_project_earnings = round(float(user_project_hours) * hourly_rate)
            earnings_cell = ws.cell(row=current_row, column=len(dates) + 4, value=user_project_earnings)
            earnings_cell.border = border
            earnings_cell.font = Font(bold=True, color="008000")  # Green color for earnings
            earnings_cell.number_format = '0 €'
            
            # Track total earnings for this user
            if user_id not in user_earnings:
                user_earnings[user_id] = 0.0
            user_earnings[user_id] += user_project_earnings
            total_earnings += user_project_earnings
            
            current_row += 1
        
        # Add project total row
        ws.cell(row=current_row, column=1, value=f"UKUPNO - {project_name}").font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = border
        
        for col, date in enumerate(dates, 3):
            date_total = sum(e.hours for e in entries if e.project_id == project_id and e.date == date)
            
            cell = ws.cell(row=current_row, column=col, value=round(date_total))
            cell.font = Font(bold=True)
            cell.border = border
            cell.number_format = '0'
        
        # Project total
        project_total_cell = ws.cell(row=current_row, column=len(dates) + 3, value=round(project_totals[project_id]))
        project_total_cell.font = Font(bold=True)
        project_total_cell.border = border
        project_total_cell.number_format = '0'
        
        current_row += 1
    
    # Add grand total row
    ws.cell(row=current_row, column=1, value=f"UKUPNO - {company.name}").font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=1).border = border
    
    grand_total = 0
    for col, date in enumerate(dates, 3):
        date_total = sum(e.hours for e in entries if e.date == date)
        grand_total += date_total
        
        cell = ws.cell(row=current_row, column=col, value=round(date_total))
        cell.font = Font(bold=True, size=12)
        cell.border = border
        cell.number_format = '0'
    
    # Grand total hours
    grand_total_cell = ws.cell(row=current_row, column=len(dates) + 3, value=round(grand_total))
    grand_total_cell.font = Font(bold=True, size=12)
    grand_total_cell.border = border
    grand_total_cell.number_format = '0'
    
    # Grand total earnings
    grand_earnings_cell = ws.cell(row=current_row, column=len(dates) + 4, value=total_earnings)
    grand_earnings_cell.font = Font(bold=True, size=12, color="008000")
    grand_earnings_cell.border = border
    grand_earnings_cell.number_format = '0 €'
    
    # Add 2 empty rows for separation
    current_row += 2
    
    # Add user summary section
    ws.cell(row=current_row, column=1, value="UKUPNO PO KORISNICIMA").font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=1).border = border
    current_row += 1
    
    # Add headers for user summary
    user_summary_headers = ['Korisnik', 'Ukupno sati', 'Ukupna zarada']
    for col, header in enumerate(user_summary_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    current_row += 1
    
    # Add user summary data
    for user_id, earnings in user_earnings.items():
        user = User.query.get(user_id)
        user_name = user.get_full_name() if user else f"Korisnik {user_id}"
        
        # Calculate total hours for this user
        user_total_hours = sum(e.hours for e in entries if e.user_id == user_id)
        
        ws.cell(row=current_row, column=1, value=user_name).border = border
        ws.cell(row=current_row, column=2, value=round(user_total_hours)).border = border
        ws.cell(row=current_row, column=2).number_format = '0'
        ws.cell(row=current_row, column=3, value=round(earnings)).border = border
        ws.cell(row=current_row, column=3).font = Font(color="008000")  # Green color for earnings
        ws.cell(row=current_row, column=3).number_format = '0 €'
        
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"kompanija_{company.name.replace(' ', '_')}{date_suffix}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/company/<int:company_id>/pdf')
@login_required
@csrf.exempt
def export_company_pdf(company_id):
    """Export detailed company report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get company details
    company = Company.query.get_or_404(company_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get time entries for the company's projects
    query = db.session.query(TimeEntry).join(Project).filter(Project.company_id == company_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(TimeEntry.date, Project.name, TimeEntry.user_id).all()
    
    # Create PDF with landscape orientation for better width
    pdf_file = BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
    
    # Register fonts for Serbian characters
    serbian_font = register_serbian_fonts()
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=serbian_font
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        spaceBefore=20,
        fontName=serbian_font
    )
    
    # Build PDF content
    story = []
    
    # Title (centered)
    story.append(Paragraph(f"Izveštaj kompanije: {company.name}", title_style))
    story.append(Spacer(1, 10))
    
    # Period information (centered)
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    story.append(Paragraph(period_text, title_style))
    story.append(Spacer(1, 20))
    
    # Company info table
    company_info = [
        ['Naziv:', company.name],
        ['Email:', company.email or 'Nije definisan'],
        ['Telefon:', company.phone or 'Nije definisan'],
        ['Website:', company.website or 'Nije definisan'],
        ['Adresa:', company.address or 'Nije definisan'],
        ['Opis:', company.description or 'Nema opisa']
    ]
    info_table = Table(company_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), serbian_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # Group entries by project for summary
    projects = {}
    for entry in entries:
        if entry.project_id not in projects:
            projects[entry.project_id] = {
                'name': entry.project.name,
                'description': entry.project.description,
                'status': entry.project.status,
                'users': {},
                'total_hours': 0
            }
        if entry.user_id not in projects[entry.project_id]['users']:
            projects[entry.project_id]['users'][entry.user_id] = {
                'name': entry.user.get_full_name(),
                'total_hours': 0
            }
        projects[entry.project_id]['users'][entry.user_id]['total_hours'] += entry.hours
        projects[entry.project_id]['total_hours'] += entry.hours
    
    # Create summary for each project
    for project_id, project_info in projects.items():
        story.append(PageBreak())
        
        # Project summary (centered)
        story.append(Paragraph(f"Projekat: {project_info['name']}", title_style))
        story.append(Spacer(1, 10))
        
        # Project basic info
        project_basic_info = [
            ['Opis:', project_info['description'] or 'Nema opisa'],
            ['Status:', project_info['status']],
            ['Ukupno sati:', f"{project_info['total_hours']:.2f}"],
            ['Korisnika:', str(len(project_info['users']))]
        ]
        basic_table = Table(project_basic_info, colWidths=[2*inch, 4*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), serbian_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (1, 0), (1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(basic_table)
        story.append(Spacer(1, 15))
        
        # Users summary for this project
        users_data = [['Korisnik', 'Ukupno sati']]
        project_total_hours = 0
        for user in project_info['users'].values():
            users_data.append([user['name'], f"{user['total_hours']:.2f}"])
            project_total_hours += user['total_hours']
        
        # Add total row
        users_data.append(['UKUPNO', f"{project_total_hours:.2f}"])
        
        users_table = Table(users_data, colWidths=[3*inch, 2*inch])
        users_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), serbian_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), serbian_font),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(Paragraph("Korisnici na projektu", title_style))
        story.append(Spacer(1, 10))
        story.append(users_table)
    
    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    
    # Create filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"kompanija_{company.name.replace(' ', '_')}{date_suffix}.pdf"
    return send_file(
        pdf_file,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/all-companies/zip')
@login_required
@csrf.exempt
def export_all_companies_zip():
    """Export all companies to ZIP with individual Excel and PDF files"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get all companies
    companies = Company.query.all()
    
    # Create ZIP file
    zip_file = BytesIO()
    with zipfile.ZipFile(zip_file, 'w') as zipf:
        for company in companies:
            # Get time entries for this company
            query = db.session.query(TimeEntry).join(Project).filter(Project.company_id == company.id)
            if start_date:
                query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
            if end_date:
                query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
            
            entries = query.order_by(TimeEntry.date, Project.name, TimeEntry.user_id).all()
            
            if not entries:
                continue  # Skip companies with no entries
            
            # Create Excel file for this company
            wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
            ws = wb.active
            ws.title = f"Kompanija - {company.name[:20]}"
            
            # Format worksheet
            start_row = format_worksheet(ws, f"Izveštaj kompanije: {company.name}")
            
            # Add period information
            period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
            ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
            
            # Get unique dates and users
            dates = sorted(list(set(entry.date for entry in entries)))
            users = {}
            for entry in entries:
                if entry.user_id not in users:
                    users[entry.user_id] = entry.user.get_full_name()
            
            # Create headers (dates as rows, users as columns)
            headers = ['Datum'] + [user_name for user_name in users.values()] + ['UKUPNO']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Fill data
            current_row = start_row + 1
            user_totals = {user_id: 0 for user_id in users.keys()}
            
            for date in dates:
                ws.cell(row=current_row, column=1, value=format_date_for_display(date)).border = border
                date_total = 0
                
                for col, (user_id, user_name) in enumerate(users.items(), 2):
                    entry = next((e for e in entries if e.user_id == user_id and e.date == date), None)
                    hours = entry.hours if entry else 0
                    user_totals[user_id] += hours
                    date_total += hours
                    
                    cell = ws.cell(row=current_row, column=col, value=hours)
                    cell.border = border
                    if hours > 0:
                        cell.number_format = '0.00'
                
                total_cell = ws.cell(row=current_row, column=len(users) + 2, value=date_total)
                total_cell.border = border
                total_cell.font = Font(bold=True)
                total_cell.number_format = '0.00'
                
                current_row += 1
            
            # Add grand total row
            ws.cell(row=current_row, column=1, value="UKUPNO").font = Font(bold=True)
            ws.cell(row=current_row, column=1).border = border
            
            grand_total = 0
            for col, (user_id, user_name) in enumerate(users.items(), 2):
                user_total = user_totals[user_id]
                grand_total += user_total
                
                cell = ws.cell(row=current_row, column=col, value=user_total)
                cell.font = Font(bold=True)
                cell.border = border
                cell.number_format = '0.00'
            
            grand_total_cell = ws.cell(row=current_row, column=len(users) + 2, value=grand_total)
            grand_total_cell.font = Font(bold=True)
            grand_total_cell.border = border
            grand_total_cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Create PDF for this company
            pdf_file = BytesIO()
            doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
            
            # Register fonts for Serbian characters
            serbian_font = register_serbian_fonts()
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName=serbian_font
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                spaceBefore=20,
                fontName=serbian_font
            )
            
            # Build PDF content
            story = []
            
            # Title (centered)
            story.append(Paragraph(f"Izveštaj kompanije: {company.name}", title_style))
            story.append(Spacer(1, 10))
            
            # Period information (centered)
            period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
            story.append(Paragraph(period_text, title_style))
            story.append(Spacer(1, 20))
            
            # Company info table
            company_info = [
                ['Naziv:', company.name],
                ['Email:', company.email or 'Nije definisan'],
                ['Telefon:', company.phone or 'Nije definisan'],
                ['Website:', company.website or 'Nije definisan'],
                ['Adresa:', company.address or 'Nije definisan'],
                ['Opis:', company.description or 'Nema opisa']
            ]
            info_table = Table(company_info, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.grey),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), serbian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(info_table)
            story.append(Spacer(1, 30))
            
            # Group entries by project for summary
            projects = {}
            for entry in entries:
                if entry.project_id not in projects:
                    projects[entry.project_id] = {
                        'name': entry.project.name,
                        'description': entry.project.description,
                        'status': entry.project.status,
                        'users': {},
                        'total_hours': 0
                    }
                if entry.user_id not in projects[entry.project_id]['users']:
                    projects[entry.project_id]['users'][entry.user_id] = {
                        'name': entry.user.get_full_name(),
                        'total_hours': 0
                    }
                projects[entry.project_id]['users'][entry.user_id]['total_hours'] += entry.hours
                projects[entry.project_id]['total_hours'] += entry.hours
            
            # Create summary for each project
            for project_id, project_info in projects.items():
                story.append(PageBreak())
                
                # Project summary (centered)
                story.append(Paragraph(f"Projekat: {project_info['name']}", title_style))
                story.append(Spacer(1, 10))
                
                # Project basic info
                project_basic_info = [
                    ['Opis:', project_info['description'] or 'Nema opisa'],
                    ['Status:', project_info['status']],
                    ['Ukupno sati:', f"{project_info['total_hours']:.2f}"],
                    ['Korisnika:', str(len(project_info['users']))]
                ]
                basic_table = Table(project_basic_info, colWidths=[2*inch, 4*inch])
                basic_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), serbian_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (1, 0), (1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(basic_table)
                story.append(Spacer(1, 15))
                
                # Users summary for this project
                users_data = [['Korisnik', 'Ukupno sati']]
                project_total_hours = 0
                for user in project_info['users'].values():
                    users_data.append([user['name'], f"{user['total_hours']:.2f}"])
                    project_total_hours += user['total_hours']
                
                # Add total row
                users_data.append(['UKUPNO', f"{project_total_hours:.2f}"])
                
                users_table = Table(users_data, colWidths=[3*inch, 2*inch])
                users_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), serbian_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
                    ('FONTNAME', (0, -1), (-1, -1), serbian_font),
                    ('FONTSIZE', (0, -1), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(Paragraph("Korisnici na projektu", title_style))
                story.append(Spacer(1, 10))
                story.append(users_table)
            
            # Build PDF
            doc.build(story)
            pdf_file.seek(0)
            
            # Add files to ZIP with date range in filename
            date_suffix = ""
            if start_date or end_date:
                date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
            
            safe_company_name = company.name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            excel_filename = f"kompanija_{safe_company_name}{date_suffix}.xlsx"
            pdf_filename = f"kompanija_{safe_company_name}{date_suffix}.pdf"
            
            zipf.writestr(excel_filename, excel_file.getvalue())
            zipf.writestr(pdf_filename, pdf_file.getvalue())
    
    zip_file.seek(0)
    
    # Create ZIP filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"sve_kompanije{date_suffix}.zip"
    
    return send_file(
        zip_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/all-projects')
@login_required
@csrf.exempt
def export_all_projects_excel():
    """Export all projects report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get all projects with time entries
    query = db.session.query(TimeEntry).join(Project).join(Company)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(Company.name, Project.name, TimeEntry.date, TimeEntry.user_id).all()
    
    # Create Excel workbook
    wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
    ws = wb.active
    ws.title = "Svi projekti"
    
    # Format worksheet
    start_row = format_worksheet(ws, "Izveštaj svih projekata")
    
    # Add period information
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
    
    # Get unique dates and projects
    dates = sorted(list(set(entry.date for entry in entries)))
    projects = {}
    for entry in entries:
        if entry.project_id not in projects:
            projects[entry.project_id] = {
                'name': entry.project.name,
                'company': entry.project.company.name
            }
    
    # Create headers
    headers = ['Kompanija', 'Projekat', 'Korisnik'] + [format_date_for_display(date) for date in dates] + ['Ukupno']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data
    current_row = start_row + 1
    project_totals = {}
    company_totals = {}
    
    for project_id, project_info in projects.items():
        project_totals[project_id] = 0
        
        # Get users for this project
        project_users = {}
        for entry in entries:
            if entry.project_id == project_id and entry.user_id not in project_users:
                project_users[entry.user_id] = entry.user.get_full_name()
        
        for user_id, user_name in project_users.items():
            ws.cell(row=current_row, column=1, value=project_info['company']).border = border
            ws.cell(row=current_row, column=2, value=project_info['name']).border = border
            ws.cell(row=current_row, column=3, value=user_name).border = border
            
            for col, date in enumerate(dates, 4):
                # Find entry for this user, project and date
                entry = next((e for e in entries if e.user_id == user_id and e.project_id == project_id and e.date == date), None)
                hours = entry.hours if entry else 0
                project_totals[project_id] += hours
                
                cell = ws.cell(row=current_row, column=col, value=hours)
                cell.border = border
                if hours > 0:
                    cell.number_format = '0.00'
            
            # Add user total for this project
            user_project_total = sum(e.hours for e in entries if e.user_id == user_id and e.project_id == project_id)
            total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=user_project_total)
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.number_format = '0.00'
            
            current_row += 1
        
        # Add project total row
        ws.cell(row=current_row, column=1, value=project_info['company']).border = border
        ws.cell(row=current_row, column=2, value=f"UKUPNO - {project_info['name']}").font = Font(bold=True)
        ws.cell(row=current_row, column=2).border = border
        
        for col, date in enumerate(dates, 4):
            date_total = sum(e.hours for e in entries if e.project_id == project_id and e.date == date)
            
            cell = ws.cell(row=current_row, column=col, value=date_total)
            cell.font = Font(bold=True)
            cell.border = border
            cell.number_format = '0.00'
        
        # Project total
        project_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=project_totals[project_id])
        project_total_cell.font = Font(bold=True)
        project_total_cell.border = border
        project_total_cell.number_format = '0.00'
        
        current_row += 1
    
    # Add grand total row
    ws.cell(row=current_row, column=1, value="UKUPNO SVI PROJEKTI").font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=1).border = border
    
    grand_total = 0
    for col, date in enumerate(dates, 4):
        date_total = sum(e.hours for e in entries if e.date == date)
        grand_total += date_total
        
        cell = ws.cell(row=current_row, column=col, value=date_total)
        cell.font = Font(bold=True, size=12)
        cell.border = border
        cell.number_format = '0.00'
    
    # Grand total
    grand_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=grand_total)
    grand_total_cell.font = Font(bold=True, size=12)
    grand_total_cell.border = border
    grand_total_cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"svi_projekti_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/user/<int:user_id>')
@login_required
@csrf.exempt
def export_user_excel(user_id):
    """Export detailed user report to Excel in the same format as personal report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get user details
    user = User.query.get_or_404(user_id)
    
    # Check permissions - users can only export their own data, admins can export any user
    if not (current_user.is_super_admin() or current_user.is_company_admin() or current_user.id == user_id):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get user's time entries with project and company information
    sql_query = """
        SELECT
            te.date,
            te.hours,
            te.description,
            p.name as project_name,
            p.id as project_id,
            c.name as company_name,
            c.id as company_id
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        JOIN companies c ON p.company_id = c.id
        WHERE te.user_id = :user_id
    """
    params = {'user_id': user_id}

    if start_date:
        sql_query += " AND te.date >= :start_date"
        params['start_date'] = parse_date_from_input(start_date)

    if end_date:
        sql_query += " AND te.date <= :end_date"
        params['end_date'] = parse_date_from_input(end_date)

    sql_query += " ORDER BY c.name, p.name, te.date"

    result = db.session.execute(text(sql_query), params)
    entries = result.fetchall()

    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from decimal import Decimal

    wb = Workbook()
    ws = wb.active
    ws.title = f"Izveštaj - {user.get_full_name()}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    company_header_font = Font(bold=True, color="FFFFFF")
    company_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    project_header_font = Font(bold=True, color="FFFFFF")
    project_header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = f"IZVEŠTAJ ZA KORISNIKA: {user.first_name} {user.last_name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    # Period
    period_text = "Period: "
    if start_date and end_date:
        period_text += f"{start_date} - {end_date}"
    elif start_date:
        period_text += f"od {start_date}"
    elif end_date:
        period_text += f"do {end_date}"
    else:
        period_text += "svi podaci"

    ws['A2'] = period_text
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')

    current_row = 4

    # Group data by company and project
    report_data = {}
    total_hours = 0

    for entry in entries:
        company_name = entry.company_name
        project_name = entry.project_name
        project_id = entry.project_id
        company_id = entry.company_id

        if company_name not in report_data:
            report_data[company_name] = {
                'company_id': company_id,
                'total_hours': 0,
                'projects': {}
            }

        if project_name not in report_data[company_name]['projects']:
            report_data[company_name]['projects'][project_name] = {
                'project_id': project_id,
                'total_hours': 0,
                'daily_entries': {}
            }

        # Add daily entry
        date_str = entry.date.strftime('%d.%m.%Y')
        if date_str not in report_data[company_name]['projects'][project_name]['daily_entries']:
            report_data[company_name]['projects'][project_name]['daily_entries'][date_str] = {
                'hours': 0,
                'description': entry.description or ''
            }

        report_data[company_name]['projects'][project_name]['daily_entries'][date_str]['hours'] += float(entry.hours)
        report_data[company_name]['projects'][project_name]['total_hours'] += float(entry.hours)
        report_data[company_name]['total_hours'] += float(entry.hours)

        total_hours += float(entry.hours)

    # Get all unique dates for the period
    all_dates = set()
    for company_name, company_data in report_data.items():
        for project_name, project_data in company_data['projects'].items():
            for date_str in project_data['daily_entries'].keys():
                all_dates.add(date_str)
    
    # Sort dates
    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
    
    # Write data to Excel with daily columns
    for company_name, company_data in report_data.items():
        # Company header
        ws[f'A{current_row}'] = f"KOMPANIJA: {company_name}"
        ws[f'A{current_row}'].font = company_header_font
        ws[f'A{current_row}'].fill = company_header_fill
        ws[f'A{current_row}'].border = border
        ws.merge_cells(f'A{current_row}:{get_column_letter(3 + len(sorted_dates))}{current_row}')
        current_row += 1

        for project_name, project_data in company_data['projects'].items():
            # Project header
            ws[f'A{current_row}'] = f"  PROJEKAT: {project_name}"
            ws[f'A{current_row}'].font = project_header_font
            ws[f'A{current_row}'].fill = project_header_fill
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'A{current_row}:{get_column_letter(3 + len(sorted_dates))}{current_row}')
            current_row += 1

            # Header row with dates
            ws[f'A{current_row}'] = "Kompanija"
            ws[f'B{current_row}'] = "Projekat"
            ws[f'C{current_row}'] = "Korisnik"
            
            # Add date columns
            for i, date_str in enumerate(sorted_dates):
                col_letter = get_column_letter(4 + i)
                ws[f'{col_letter}{current_row}'] = date_str
                ws[f'{col_letter}{current_row}'].font = header_font
                ws[f'{col_letter}{current_row}'].fill = header_fill
                ws[f'{col_letter}{current_row}'].border = border
            
            # Add total column
            total_col = get_column_letter(4 + len(sorted_dates))
            ws[f'{total_col}{current_row}'] = "UKUPNO"
            ws[f'{total_col}{current_row}'].font = header_font
            ws[f'{total_col}{current_row}'].fill = header_fill
            ws[f'{total_col}{current_row}'].border = border
            
            # Style header row
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = header_font
                ws[f'{col}{current_row}'].fill = header_fill
                ws[f'{col}{current_row}'].border = border
            current_row += 1

            # Data row
            ws[f'A{current_row}'] = company_name
            ws[f'B{current_row}'] = project_name
            ws[f'C{current_row}'] = f"{user.first_name} {user.last_name}"
            
            # Add hours for each date
            for i, date_str in enumerate(sorted_dates):
                col_letter = get_column_letter(4 + i)
                hours = project_data['daily_entries'].get(date_str, {}).get('hours', 0)
                ws[f'{col_letter}{current_row}'] = hours if hours > 0 else ""
                ws[f'{col_letter}{current_row}'].border = border
                if hours > 0:
                    ws[f'{col_letter}{current_row}'].number_format = '0.00'
            
            # Add project total
            total_col = get_column_letter(4 + len(sorted_dates))
            ws[f'{total_col}{current_row}'] = project_data['total_hours']
            ws[f'{total_col}{current_row}'].font = Font(bold=True)
            ws[f'{total_col}{current_row}'].border = border
            
            # Style data row
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].border = border
            current_row += 2

        # Company total
        current_row += 1
        ws[f'A{current_row}'] = f"UKUPNO ZA KOMPANIJU: {company_name}"
        ws[f'A{current_row}'].font = total_font
        ws[f'A{current_row}'].fill = total_fill
        ws[f'A{current_row}'].border = border
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1

        # Company total row
        ws[f'A{current_row}'] = "UKUPNO"
        ws[f'B{current_row}'] = company_data['total_hours']
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].font = total_font
            ws[f'{col}{current_row}'].fill = total_fill
            ws[f'{col}{current_row}'].border = border
        current_row += 3  # Extra space between companies

    # Final summary table
    current_row += 2
    ws[f'A{current_row}'] = "UKUPAN PREGLED PO KOMPANIJAMA"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 2

    # Summary table header
    ws[f'A{current_row}'] = "Kompanija"
    ws[f'B{current_row}'] = "Naziv korisnika"
    ws[f'C{current_row}'] = "Broj radnih sati"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].font = header_font
        ws[f'{col}{current_row}'].fill = header_fill
        ws[f'{col}{current_row}'].border = border
    current_row += 1

    # Summary table data
    for company_name, company_data in report_data.items():
        ws[f'A{current_row}'] = company_name
        ws[f'B{current_row}'] = f"{user.first_name} {user.last_name}"
        ws[f'C{current_row}'] = company_data['total_hours']
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = border
        current_row += 1

    # Final total row
    ws[f'A{current_row}'] = "UKUPAN ZBIR ZA PERIOD"
    ws[f'B{current_row}'] = ""
    ws[f'C{current_row}'] = total_hours
    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].font = total_font
        ws[f'{col}{current_row}'].fill = total_fill
        ws[f'{col}{current_row}'].border = border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"izvestaj_korisnika_{user.username}"
    if start_date:
        filename += f"_od_{start_date.replace('.', '-')}"
    if end_date:
        filename += f"_do_{end_date.replace('.', '-')}"
    filename += ".xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reports.route('/export/all-users/zip')
@login_required
@csrf.exempt
def export_all_users_zip():
    """Export all users to ZIP with individual Excel and PDF files"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get all users
    users = User.query.all()
    
    # Create ZIP file
    zip_file = BytesIO()
    with zipfile.ZipFile(zip_file, 'w') as zipf:
        for user in users:
            # Get time entries for this user
            query = db.session.query(TimeEntry).filter(TimeEntry.user_id == user.id)
            if start_date:
                query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
            if end_date:
                query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
            
            entries = query.order_by(TimeEntry.date, TimeEntry.project_id).all()
            
            if not entries:
                continue  # Skip users with no entries
            
            # Create Excel file for this user
            wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
            ws = wb.active
            ws.title = f"Korisnik - {user.get_full_name()[:20]}"
            
            # Format worksheet
            start_row = format_worksheet(ws, f"Izveštaj korisnika: {user.get_full_name()}")
            
            # Add period information
            period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
            ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
            
            # Get unique dates and projects
            dates = sorted(list(set(entry.date for entry in entries)))
            projects = {}
            for entry in entries:
                if entry.project_id not in projects:
                    projects[entry.project_id] = {
                        'name': entry.project.name,
                        'company': entry.project.company.name
                    }
            
            # Create headers (dates as rows, projects as columns)
            headers = ['Datum'] + [f"{project_info['name']} ({project_info['company']})" for project_info in projects.values()] + ['UKUPNO']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Fill data
            current_row = start_row + 1
            project_totals = {project_id: 0 for project_id in projects.keys()}
            
            for date in dates:
                ws.cell(row=current_row, column=1, value=format_date_for_display(date)).border = border
                date_total = 0
                
                for col, (project_id, project_info) in enumerate(projects.items(), 2):
                    entry = next((e for e in entries if e.project_id == project_id and e.date == date), None)
                    hours = entry.hours if entry else 0
                    project_totals[project_id] += hours
                    date_total += hours
                    
                    cell = ws.cell(row=current_row, column=col, value=hours)
                    cell.border = border
                    if hours > 0:
                        cell.number_format = '0.00'
                
                total_cell = ws.cell(row=current_row, column=len(projects) + 2, value=date_total)
                total_cell.border = border
                total_cell.font = Font(bold=True)
                total_cell.number_format = '0.00'
                
                current_row += 1
            
            # Add grand total row
            ws.cell(row=current_row, column=1, value="UKUPNO").font = Font(bold=True)
            ws.cell(row=current_row, column=1).border = border
            
            grand_total = 0
            for col, (project_id, project_info) in enumerate(projects.items(), 2):
                project_total = project_totals[project_id]
                grand_total += project_total
                
                cell = ws.cell(row=current_row, column=col, value=project_total)
                cell.font = Font(bold=True)
                cell.border = border
                cell.number_format = '0.00'
            
            grand_total_cell = ws.cell(row=current_row, column=len(projects) + 2, value=grand_total)
            grand_total_cell.font = Font(bold=True)
            grand_total_cell.border = border
            grand_total_cell.number_format = '0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Create PDF for this user
            pdf_file = BytesIO()
            doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
            
            # Register fonts for Serbian characters
            serbian_font = register_serbian_fonts()
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName=serbian_font
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                spaceBefore=20,
                fontName=serbian_font
            )
            
            # Build PDF content
            story = []
            
            # Title (centered)
            story.append(Paragraph(f"Izveštaj korisnika: {user.get_full_name()}", title_style))
            story.append(Spacer(1, 10))
            
            # Period information (centered)
            period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
            story.append(Paragraph(period_text, title_style))
            story.append(Spacer(1, 20))
            
            # User info table
            user_info = [
                ['Ime i prezime:', user.get_full_name()],
                ['Username:', user.username],
                ['Email:', user.email or 'Nije definisan'],
                ['Rola:', user.role],
                ['Ukupno sati:', f"{sum(e.hours for e in entries):.2f}"],
                ['Ukupno unosa:', str(len(entries))],
                ['Projekata:', str(len(set(e.project_id for e in entries)))],
                ['Kompanija:', str(len(set(e.project.company_id for e in entries)))]
            ]
            info_table = Table(user_info, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.grey),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), serbian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(info_table)
            story.append(Spacer(1, 30))
            
            # Group entries by project for summary
            projects = {}
            for entry in entries:
                if entry.project_id not in projects:
                    projects[entry.project_id] = {
                        'name': entry.project.name,
                        'company': entry.project.company.name,
                        'description': entry.project.description,
                        'status': entry.project.status,
                        'total_hours': 0,
                        'entry_count': 0
                    }
                projects[entry.project_id]['total_hours'] += entry.hours
                projects[entry.project_id]['entry_count'] += 1
            
            # Create summary for each project
            for project_id, project_info in projects.items():
                story.append(PageBreak())
                
                # Project summary (centered)
                story.append(Paragraph(f"Projekat: {project_info['name']}", title_style))
                story.append(Spacer(1, 10))
                
                # Project basic info
                project_basic_info = [
                    ['Kompanija:', project_info['company']],
                    ['Opis:', project_info['description'] or 'Nema opisa'],
                    ['Status:', project_info['status']],
                    ['Ukupno sati:', f"{project_info['total_hours']:.2f}"],
                    ['Broj unosa:', str(project_info['entry_count'])]
                ]
                basic_table = Table(project_basic_info, colWidths=[2*inch, 4*inch])
                basic_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), serbian_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (1, 0), (1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(basic_table)
                story.append(Spacer(1, 15))
            
            # Add total summary at the end
            story.append(PageBreak())
            story.append(Paragraph("Ukupan pregled", title_style))
            story.append(Spacer(1, 20))
            
            total_hours = sum(e.hours for e in entries)
            total_entries = len(entries)
            
            summary_data = [
                ['Ukupno sati:', f"{total_hours:.2f}"],
                ['Ukupno unosa:', str(total_entries)],
                ['Projekata:', str(len(projects))],
                ['Kompanija:', str(len(set(e.project.company_id for e in entries)))]
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), serbian_font),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (1, 0), (1, -1), colors.lightyellow),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            
            # Build PDF
            doc.build(story)
            pdf_file.seek(0)
            
            # Add files to ZIP with date range in filename
            date_suffix = ""
            if start_date or end_date:
                date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
            
            safe_username = user.username.replace(' ', '_').replace('/', '_').replace('\\', '_')
            excel_filename = f"korisnik_{safe_username}{date_suffix}.xlsx"
            pdf_filename = f"korisnik_{safe_username}{date_suffix}.pdf"
            
            zipf.writestr(excel_filename, excel_file.getvalue())
            zipf.writestr(pdf_filename, pdf_file.getvalue())
    
    zip_file.seek(0)
    
    # Create ZIP filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"svi_korisnici{date_suffix}.zip"
    
    return send_file(
        zip_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/user/<int:user_id>/pdf')
@login_required
@csrf.exempt
def export_user_pdf(user_id):
    """Export detailed user report to PDF"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get user details
    user = User.query.get_or_404(user_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin() or current_user.id == user_id):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get time entries for the user
    query = TimeEntry.query.filter_by(user_id=user_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(TimeEntry.date, TimeEntry.project_id).all()
    
    # Create PDF with landscape orientation for better width
    pdf_file = BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=72)
    
    # Register fonts for Serbian characters
    serbian_font = register_serbian_fonts()
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=serbian_font
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        spaceBefore=20,
        fontName=serbian_font
    )
    
    # Build PDF content
    story = []
    
    # Title (centered)
    story.append(Paragraph(f"Izveštaj korisnika: {user.get_full_name()}", title_style))
    story.append(Spacer(1, 10))
    
    # Period information (centered)
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    story.append(Paragraph(period_text, title_style))
    story.append(Spacer(1, 20))
    
    # User info table
    user_info = [
        ['Ime i prezime:', user.get_full_name()],
        ['Username:', user.username],
        ['Email:', user.email or 'Nije definisan'],
        ['Rola:', user.role],
        ['Ukupno sati:', f"{sum(e.hours for e in entries):.2f}"],
        ['Ukupno unosa:', str(len(entries))],
        ['Projekata:', str(len(set(e.project_id for e in entries)))],
        ['Kompanija:', str(len(set(e.project.company_id for e in entries)))]
    ]
    info_table = Table(user_info, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), serbian_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # Group entries by project for summary
    projects = {}
    for entry in entries:
        if entry.project_id not in projects:
            projects[entry.project_id] = {
                'name': entry.project.name,
                'company': entry.project.company.name,
                'description': entry.project.description,
                'status': entry.project.status,
                'total_hours': 0,
                'entry_count': 0
            }
        projects[entry.project_id]['total_hours'] += entry.hours
        projects[entry.project_id]['entry_count'] += 1
    
    # Create summary for each project
    for project_id, project_info in projects.items():
        story.append(PageBreak())
        
        # Project summary (centered)
        story.append(Paragraph(f"Projekat: {project_info['name']}", title_style))
        story.append(Spacer(1, 10))
        
        # Project basic info
        project_basic_info = [
            ['Kompanija:', project_info['company']],
            ['Opis:', project_info['description'] or 'Nema opisa'],
            ['Status:', project_info['status']],
            ['Ukupno sati:', f"{project_info['total_hours']:.2f}"],
            ['Broj unosa:', str(project_info['entry_count'])]
        ]
        basic_table = Table(project_basic_info, colWidths=[2*inch, 4*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), serbian_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (1, 0), (1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(basic_table)
        story.append(Spacer(1, 15))
    
    # Add total summary at the end
    story.append(PageBreak())
    story.append(Paragraph("Ukupan pregled", title_style))
    story.append(Spacer(1, 20))
    
    total_hours = sum(e.hours for e in entries)
    total_entries = len(entries)
    
    summary_data = [
        ['Ukupno sati:', f"{total_hours:.2f}"],
        ['Ukupno unosa:', str(total_entries)],
        ['Projekata:', str(len(projects))],
        ['Kompanija:', str(len(set(e.project.company_id for e in entries)))]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), serbian_font),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (1, 0), (1, -1), colors.lightyellow),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    
    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    
    # Create filename with date range
    date_suffix = ""
    if start_date or end_date:
        date_suffix = f"_{start_date or 'svi'}_{end_date or 'svi'}"
    
    filename = f"korisnik_{user.username}{date_suffix}.pdf"
    
    return send_file(
        pdf_file,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
    
    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    
    filename = f"korisnik_{user.username}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    
    return send_file(
        pdf_file,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/all-users')
@login_required
@csrf.exempt
def export_all_users_excel():
    """Export all users report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get all users with time entries
    query = db.session.query(TimeEntry).join(User).join(Project).join(Company)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(User.last_name, User.first_name, TimeEntry.date, TimeEntry.project_id).all()
    
    # Create Excel workbook
    wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
    ws = wb.active
    ws.title = "Svi korisnici"
    
    # Format worksheet
    start_row = format_worksheet(ws, "Izveštaj svih korisnika")
    
    # Add period information
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
    
    # Get unique dates and users
    dates = sorted(list(set(entry.date for entry in entries)))
    users = {}
    for entry in entries:
        if entry.user_id not in users:
            users[entry.user_id] = {
                'name': entry.user.get_full_name(),
                'username': entry.user.username,
                'projects': {}
            }
        if entry.project_id not in users[entry.user_id]['projects']:
            users[entry.user_id]['projects'][entry.project_id] = {
                'name': entry.project.name,
                'company': entry.project.company.name
            }
    
    # Create headers
    headers = ['Korisnik', 'Projekat', 'Kompanija'] + [format_date_for_display(date) for date in dates] + ['Ukupno']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data
    current_row = start_row + 1
    user_totals = {}
    project_totals = {}
    
    for user_id, user_info in users.items():
        user_totals[user_id] = 0
        
        for project_id, project_info in user_info['projects'].items():
            project_totals[project_id] = 0
            
            ws.cell(row=current_row, column=1, value=user_info['name']).border = border
            ws.cell(row=current_row, column=2, value=project_info['name']).border = border
            ws.cell(row=current_row, column=3, value=project_info['company']).border = border
            
            for col, date in enumerate(dates, 4):
                # Find entry for this user, project and date
                entry = next((e for e in entries if e.user_id == user_id and e.project_id == project_id and e.date == date), None)
                hours = entry.hours if entry else 0
                project_totals[project_id] += hours
                user_totals[user_id] += hours
                
                cell = ws.cell(row=current_row, column=col, value=hours)
                cell.border = border
                if hours > 0:
                    cell.number_format = '0.00'
            
            # Add user total for this project
            user_project_total = sum(e.hours for e in entries if e.user_id == user_id and e.project_id == project_id)
            total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=user_project_total)
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.number_format = '0.00'
            
            current_row += 1
        
        # Add user total row
        ws.cell(row=current_row, column=1, value=f"UKUPNO - {user_info['name']}").font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = border
        
        for col, date in enumerate(dates, 4):
            date_total = sum(e.hours for e in entries if e.user_id == user_id and e.date == date)
            
            cell = ws.cell(row=current_row, column=col, value=date_total)
            cell.font = Font(bold=True)
            cell.border = border
            cell.number_format = '0.00'
        
        # User total
        user_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=user_totals[user_id])
        user_total_cell.font = Font(bold=True)
        user_total_cell.border = border
        user_total_cell.number_format = '0.00'
        
        current_row += 1
    
    # Add grand total row
    ws.cell(row=current_row, column=1, value="UKUPNO SVI KORISNICI").font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=1).border = border
    
    grand_total = 0
    for col, date in enumerate(dates, 4):
        date_total = sum(e.hours for e in entries if e.date == date)
        grand_total += date_total
        
        cell = ws.cell(row=current_row, column=col, value=date_total)
        cell.font = Font(bold=True, size=14)
        cell.border = border
        cell.number_format = '0.00'
    
    # Grand total
    grand_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=grand_total)
    grand_total_cell.font = Font(bold=True, size=14)
    grand_total_cell.border = border
    grand_total_cell.number_format = '0.00'
    
    # Add summary table
    current_row += 3  # Add some space
    
    # Summary table title
    ws.cell(row=current_row, column=1, value="UKUPAN PREGLED PO KOMPANIJAMA, PROJEKTIMA I KORISNICIMA").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    # Summary table header
    summary_headers = ['Kompanija', 'Projekat', 'Korisnik', 'Ukupno sati']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    current_row += 1
    
    # Group data for summary
    summary_data = {}
    for entry in entries:
        company_name = entry.project.company.name
        project_name = entry.project.name
        user_name = entry.user.get_full_name()
        
        key = (company_name, project_name, user_name)
        if key not in summary_data:
            summary_data[key] = 0
        summary_data[key] += float(entry.hours)
    
    # Sort summary data by company, then project, then user
    sorted_summary = sorted(summary_data.items(), key=lambda x: (x[0][0], x[0][1], x[0][2]))
    
    # Add summary data
    company_totals = {}
    project_totals = {}
    user_totals = {}
    
    for (company_name, project_name, user_name), hours in sorted_summary:
        ws.cell(row=current_row, column=1, value=company_name).border = border
        ws.cell(row=current_row, column=2, value=project_name).border = border
        ws.cell(row=current_row, column=3, value=user_name).border = border
        ws.cell(row=current_row, column=4, value=hours).border = border
        ws.cell(row=current_row, column=4).number_format = '0.00'
        
        # Track totals
        if company_name not in company_totals:
            company_totals[company_name] = 0
        company_totals[company_name] += hours
        
        project_key = (company_name, project_name)
        if project_key not in project_totals:
            project_totals[project_key] = 0
        project_totals[project_key] += hours
        
        if user_name not in user_totals:
            user_totals[user_name] = 0
        user_totals[user_name] += hours
        
        current_row += 1
    
    # Add company totals
    current_row += 2
    ws.cell(row=current_row, column=1, value="UKUPNO PO KOMPANIJAMA").font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    
    for company_name, total_hours in sorted(company_totals.items()):
        ws.cell(row=current_row, column=1, value=company_name).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=4, value=total_hours).font = Font(bold=True)
        ws.cell(row=current_row, column=4).border = border
        ws.cell(row=current_row, column=4).number_format = '0.00'
        current_row += 1
    
    # Add project totals
    current_row += 2
    ws.cell(row=current_row, column=1, value="UKUPNO PO PROJEKTIMA").font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    
    for (company_name, project_name), total_hours in sorted(project_totals.items()):
        ws.cell(row=current_row, column=1, value=company_name).border = border
        ws.cell(row=current_row, column=2, value=project_name).font = Font(bold=True)
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=4, value=total_hours).font = Font(bold=True)
        ws.cell(row=current_row, column=4).border = border
        ws.cell(row=current_row, column=4).number_format = '0.00'
        current_row += 1
    
    # Add user totals
    current_row += 2
    ws.cell(row=current_row, column=1, value="UKUPNO PO KORISNICIMA").font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1
    
    for user_name, total_hours in sorted(user_totals.items()):
        ws.cell(row=current_row, column=3, value=user_name).font = Font(bold=True)
        ws.cell(row=current_row, column=3).border = border
        ws.cell(row=current_row, column=4, value=total_hours).font = Font(bold=True)
        ws.cell(row=current_row, column=4).border = border
        ws.cell(row=current_row, column=4).number_format = '0.00'
        current_row += 1
    
    # Final grand total
    current_row += 2
    ws.cell(row=current_row, column=1, value="UKUPAN ZBIR ZA PERIOD").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws.cell(row=current_row, column=4, value=grand_total).font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=4).number_format = '0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"svi_korisnici_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/export/all-companies')
@login_required
@csrf.exempt
def export_all_companies_excel():
    """Export all companies report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get all companies with time entries
    query = db.session.query(TimeEntry).join(Project).join(Company)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(Company.name, Project.name, TimeEntry.date, TimeEntry.user_id).all()
    
    # Create Excel workbook
    wb, header_font, header_fill, header_alignment, border = create_excel_workbook()
    ws = wb.active
    ws.title = "Sve kompanije"
    
    # Format worksheet
    start_row = format_worksheet(ws, "Izveštaj svih kompanija")
    
    # Add period information
    period_text = f"Period izveštaja: {start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    ws.cell(row=start_row - 1, column=1, value=period_text).font = Font(bold=True, italic=True)
    
    # Get unique dates and companies
    dates = sorted(list(set(entry.date for entry in entries)))
    companies = {}
    for entry in entries:
        if entry.project.company_id not in companies:
            companies[entry.project.company_id] = {
                'name': entry.project.company.name,
                'projects': {}
            }
        if entry.project_id not in companies[entry.project.company_id]['projects']:
            companies[entry.project.company_id]['projects'][entry.project_id] = entry.project.name
    
    # Create headers
    headers = ['Kompanija', 'Projekat', 'Korisnik'] + [format_date_for_display(date) for date in dates] + ['Ukupno', 'Zarada']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data
    current_row = start_row + 1
    company_totals = {}
    project_totals = {}
    user_earnings = {}  # Track earnings per user
    total_earnings = 0  # Track total earnings
    
    for company_id, company_info in companies.items():
        company_totals[company_id] = 0
        
        for project_id, project_name in company_info['projects'].items():
            project_totals[project_id] = 0
            
            # Get users for this project
            project_users = {}
            for entry in entries:
                if entry.project_id == project_id and entry.user_id not in project_users:
                    project_users[entry.user_id] = entry.user.get_full_name()
            
            for user_id, user_name in project_users.items():
                # Get user's hourly rate
                user = User.query.get(user_id)
                hourly_rate = float(user.hourly_rate) if user and user.hourly_rate else 0.0
                
                ws.cell(row=current_row, column=1, value=company_info['name']).border = border
                ws.cell(row=current_row, column=2, value=project_name).border = border
                ws.cell(row=current_row, column=3, value=user_name).border = border
                
                user_project_hours = 0
                for col, date in enumerate(dates, 4):
                    # Find entry for this user, project and date
                    entry = next((e for e in entries if e.user_id == user_id and e.project_id == project_id and e.date == date), None)
                    hours = entry.hours if entry else 0
                    user_project_hours += hours
                    project_totals[project_id] += hours
                    company_totals[company_id] += hours
                    
                    cell = ws.cell(row=current_row, column=col, value=round(hours) if hours > 0 else 0)
                    cell.border = border
                    if hours > 0:
                        cell.number_format = '0'
                
                # Add user total for this project
                total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=round(user_project_hours))
                total_cell.border = border
                total_cell.font = Font(bold=True)
                total_cell.number_format = '0'
                
                # Calculate and add user earnings for this project
                user_project_earnings = round(float(user_project_hours) * hourly_rate)
                earnings_cell = ws.cell(row=current_row, column=len(dates) + 5, value=user_project_earnings)
                earnings_cell.border = border
                earnings_cell.font = Font(bold=True, color="008000")  # Green color for earnings
                earnings_cell.number_format = '0 €'
                
                # Track total earnings for this user
                if user_id not in user_earnings:
                    user_earnings[user_id] = 0.0
                user_earnings[user_id] += user_project_earnings
                total_earnings += user_project_earnings
                
                current_row += 1
            
            # Add project total row
            ws.cell(row=current_row, column=1, value=company_info['name']).border = border
            ws.cell(row=current_row, column=2, value=f"UKUPNO - {project_name}").font = Font(bold=True)
            ws.cell(row=current_row, column=2).border = border
            
            for col, date in enumerate(dates, 4):
                date_total = sum(e.hours for e in entries if e.project_id == project_id and e.date == date)
                
                cell = ws.cell(row=current_row, column=col, value=round(date_total))
                cell.font = Font(bold=True)
                cell.border = border
                cell.number_format = '0'
            
            # Project total
            project_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=round(project_totals[project_id]))
            project_total_cell.font = Font(bold=True)
            project_total_cell.border = border
            project_total_cell.number_format = '0'
            
            current_row += 1
        
        # Add 2 empty rows between companies for better separation
        current_row += 2
        
        # Add company total row
        ws.cell(row=current_row, column=1, value=f"UKUPNO - {company_info['name']}").font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).border = border
        
        for col, date in enumerate(dates, 4):
            date_total = sum(e.hours for e in entries if e.project.company_id == company_id and e.date == date)
            
            cell = ws.cell(row=current_row, column=col, value=round(date_total))
            cell.font = Font(bold=True, size=12)
            cell.border = border
            cell.number_format = '0'
        
        # Company total
        company_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=round(company_totals[company_id]))
        company_total_cell.font = Font(bold=True, size=12)
        company_total_cell.border = border
        company_total_cell.number_format = '0'
        
        current_row += 1
    
    # Add grand total row
    ws.cell(row=current_row, column=1, value="UKUPNO SVE KOMPANIJE").font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=1).border = border
    
    grand_total = 0
    for col, date in enumerate(dates, 4):
        date_total = sum(e.hours for e in entries if e.date == date)
        grand_total += date_total
        
        cell = ws.cell(row=current_row, column=col, value=round(date_total))
        cell.font = Font(bold=True, size=14)
        cell.border = border
        cell.number_format = '0'
    
    # Grand total hours
    grand_total_cell = ws.cell(row=current_row, column=len(dates) + 4, value=round(grand_total))
    grand_total_cell.font = Font(bold=True, size=14)
    grand_total_cell.border = border
    grand_total_cell.number_format = '0'
    
    # Grand total earnings
    grand_earnings_cell = ws.cell(row=current_row, column=len(dates) + 5, value=total_earnings)
    grand_earnings_cell.font = Font(bold=True, size=14, color="008000")
    grand_earnings_cell.border = border
    grand_earnings_cell.number_format = '0 €'
    
    # Add 2 empty rows for separation
    current_row += 2
    
    # Add user summary section
    ws.cell(row=current_row, column=1, value="UKUPNO PO KORISNICIMA").font = Font(bold=True, size=14)
    ws.cell(row=current_row, column=1).border = border
    current_row += 1
    
    # Add headers for user summary
    user_summary_headers = ['Korisnik', 'Ukupno sati', 'Ukupna zarada']
    for col, header in enumerate(user_summary_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    current_row += 1
    
    # Add user summary data
    for user_id, earnings in user_earnings.items():
        user = User.query.get(user_id)
        user_name = user.get_full_name() if user else f"Korisnik {user_id}"
        
        # Calculate total hours for this user
        user_total_hours = sum(e.hours for e in entries if e.user_id == user_id)
        
        ws.cell(row=current_row, column=1, value=user_name).border = border
        ws.cell(row=current_row, column=2, value=round(user_total_hours)).border = border
        ws.cell(row=current_row, column=2).number_format = '0'
        ws.cell(row=current_row, column=3, value=round(earnings)).border = border
        ws.cell(row=current_row, column=3).font = Font(color="008000")  # Green color for earnings
        ws.cell(row=current_row, column=3).number_format = '0 €'
        
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"sve_kompanije_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports.route('/api/report-data')
@login_required
@csrf.exempt
def api_report_data():
    """API endpoint for chart data"""
    report_type = request.args.get('type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = TimeEntry.query
    
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        query = query.filter_by(user_id=current_user.id)
    
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    if report_type == 'daily':
        results = db.session.execute(text("""
            SELECT te.date, SUM(te.hours) as total_hours
            FROM time_entries te
            WHERE 1=1
            """ + (" AND te.user_id = :user_id" if not (current_user.is_super_admin() or current_user.is_company_admin()) else "") + """
            """ + (" AND te.date >= :start_date" if start_date else "") + """
            """ + (" AND te.date <= :end_date" if end_date else "") + """
            GROUP BY te.date
            ORDER BY te.date
        """), {
            'user_id': current_user.id if not (current_user.is_super_admin() or current_user.is_company_admin()) else None,
            'start_date': parse_date_from_input(start_date) if start_date else None,
            'end_date': parse_date_from_input(end_date) if end_date else None
        })
        
        return jsonify([{
            'date': format_date_for_api(result.date),
            'total_hours': float(result.total_hours)
        } for result in results])
    
    elif report_type == 'user_summary_stats':
        results = db.session.execute(text("""
            SELECT 
                CONCAT(u.first_name, ' ', u.last_name) as user_name,
                SUM(te.hours) as total_hours
            FROM time_entries te
            JOIN users u ON te.user_id = u.id
            WHERE 1=1
            """ + (" AND te.user_id = :user_id" if not (current_user.is_super_admin() or current_user.is_company_admin()) else "") + """
            """ + (" AND te.date >= :start_date" if start_date else "") + """
            """ + (" AND te.date <= :end_date" if end_date else "") + """
            GROUP BY u.id
            ORDER BY total_hours DESC
        """), {
            'user_id': current_user.id if not (current_user.is_super_admin() or current_user.is_company_admin()) else None,
            'start_date': parse_date_from_input(start_date) if start_date else None,
            'end_date': parse_date_from_input(end_date) if end_date else None
        })
        
        return jsonify([{
            'user_name': result.user_name,
            'total_hours': float(result.total_hours)
        } for result in results])
    
    elif report_type == 'stats':
        # Get quick statistics
        total_hours = db.session.query(func.sum(TimeEntry.hours)).filter(query).scalar() or 0
        
        # Get active projects count
        active_projects = db.session.query(func.count(func.distinct(TimeEntry.project_id))).filter(query).scalar() or 0
        
        # Get active users count
        active_users = db.session.query(func.count(func.distinct(TimeEntry.user_id))).filter(query).scalar() or 0
        
        # Calculate average hours per day
        if start_date and end_date:
            start = parse_date_from_input(start_date)
            end = parse_date_from_input(end_date)
            days_diff = (end - start).days + 1
            avg_hours = total_hours / days_diff if days_diff > 0 else 0
        else:
            avg_hours = 0
        
        return jsonify({
            'total_hours': round(float(total_hours), 2),
            'active_projects': active_projects,
            'active_users': active_users,
            'avg_hours': round(float(avg_hours), 2)
        })
    
    elif report_type == 'user_project':
        # Get user-project summary for pie chart
        results = db.session.execute(text("""
            SELECT 
                p.name as project_name,
                SUM(te.hours) as total_hours
            FROM time_entries te
            JOIN projects p ON te.project_id = p.id
            WHERE 1=1
            """ + (" AND te.user_id = :user_id" if not (current_user.is_super_admin() or current_user.is_company_admin()) else "") + """
            """ + (" AND te.date >= :start_date" if start_date else "") + """
            """ + (" AND te.date <= :end_date" if end_date else "") + """
            GROUP BY p.id
            ORDER BY total_hours DESC
        """), {
            'user_id': current_user.id if not (current_user.is_super_admin() or current_user.is_company_admin()) else None,
            'start_date': parse_date_from_input(start_date) if start_date else None,
            'end_date': parse_date_from_input(end_date) if end_date else None
        })
        
        return jsonify([{
            'project_name': result.project_name,
            'total_hours': float(result.total_hours)
        } for result in results])
    
    return jsonify([])

@reports.route('/api/project-details/<int:project_id>')
@login_required
@csrf.exempt
def api_project_details(project_id):
    """Get detailed project information"""
    project = Project.query.get_or_404(project_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin() or 
            current_user.is_project_admin(project_id)):
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query with filters
    query = TimeEntry.query.filter_by(project_id=project_id)
    
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    # Get filtered statistics
    total_hours = db.session.query(func.sum(TimeEntry.hours)).filter(query.whereclause).scalar() or 0
    total_entries = db.session.query(func.count(TimeEntry.id)).filter(query.whereclause).scalar() or 0
    
    # Get users working on this project with filters
    user_query = text("""
        SELECT u.first_name, u.last_name, SUM(te.hours) as total_hours
        FROM time_entries te
        JOIN users u ON te.user_id = u.id
        WHERE te.project_id = :project_id
        """ + (" AND te.date >= :start_date" if start_date else "") + """
        """ + (" AND te.date <= :end_date" if end_date else "") + """
        GROUP BY u.id
        ORDER BY total_hours DESC
    """)
    
    params = {'project_id': project_id}
    if start_date:
        params['start_date'] = parse_date_from_input(start_date)
    if end_date:
        params['end_date'] = parse_date_from_input(end_date)
    
    result = db.session.execute(user_query, params)
    
    users = [{
        'name': f"{row.first_name} {row.last_name}",
        'total_hours': float(row.total_hours)
    } for row in result.fetchall()]
    
    return jsonify({
        'project': {
            'id': project.id,
            'name': project.name,
            'description': project.description,
            'company_name': project.company.name,
            'start_date': format_date_for_api(project.start_date),
            'end_date': format_date_for_api(project.end_date) if project.end_date else None,
            'status': project.status,
            'budget': float(project.budget) if project.budget else None
        },
        'statistics': {
            'total_hours': float(total_hours),
            'total_entries': total_entries,
            'users': users
        },
        'report_period': {
            'start_date': start_date,
            'end_date': end_date
        }
    })

@reports.route('/api/company-details/<int:company_id>')
@login_required
@csrf.exempt
def api_company_details(company_id):
    """Get detailed company information"""
    company = Company.query.get_or_404(company_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get company statistics
    result = db.session.execute(text("""
        SELECT 
            p.name as project_name,
            SUM(te.hours) as total_hours,
            COUNT(te.id) as total_entries
        FROM time_entries te
        JOIN projects p ON te.project_id = p.id
        WHERE p.company_id = :company_id
        GROUP BY p.id
        ORDER BY total_hours DESC
    """), {'company_id': company_id})
    
    projects = [{
        'name': row.project_name,
        'total_hours': float(row.total_hours),
        'total_entries': row.total_entries
    } for row in result.fetchall()]
    
    total_hours = sum(p['total_hours'] for p in projects)
    total_entries = sum(p['total_entries'] for p in projects)
    
    return jsonify({
        'company': {
            'id': company.id,
            'name': company.name,
            'email': company.email,
            'phone': company.phone,
            'website': company.website,
            'address': company.address,
            'description': company.description
        },
        'statistics': {
            'total_hours': total_hours,
            'total_entries': total_entries,
            'projects': projects
        }
    })

@reports.route('/api/user-details/<int:user_id>')
@login_required
@csrf.exempt
def api_user_details(user_id):
    """API endpoint to get detailed user information"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get user details
    user = User.query.get_or_404(user_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin() or current_user.id == user_id):
        return jsonify({'error': 'Nemate dozvolu za pristup ovim podacima'}), 403
    
    # Get time entries for the user
    query = TimeEntry.query.filter_by(user_id=user_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(TimeEntry.date, TimeEntry.project_id).all()
    
    # Calculate statistics
    total_hours = sum(e.hours for e in entries)
    entry_count = len(entries)
    project_count = len(set(e.project_id for e in entries))
    company_count = len(set(e.project.company_id for e in entries))
    
    # Get projects summary
    projects = {}
    for entry in entries:
        if entry.project_id not in projects:
            projects[entry.project_id] = {
                'name': entry.project.name,
                'company': entry.project.company.name,
                'hours': 0,
                'entries': 0
            }
        projects[entry.project_id]['hours'] += entry.hours
        projects[entry.project_id]['entries'] += 1
    
    # Create report period text
    report_period = f"{start_date or 'Svi datumi'} - {end_date or 'Svi datumi'}"
    
    return jsonify({
        'user_id': user.id,
        'user_name': user.get_full_name(),
        'username': user.username,
        'email': user.email,
        'role': user.role,
        'total_hours': total_hours,
        'entry_count': entry_count,
        'project_count': project_count,
        'company_count': company_count,
        'report_period': report_period,
        'projects': list(projects.values())
    }) 

@reports.route('/company/<int:company_id>')
@login_required
def company_detail(company_id):
    """Detailed view for a single company"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get company details
    company = Company.query.get_or_404(company_id)
    
    # Check permissions
    if not (current_user.is_super_admin() or current_user.is_company_admin()):
        flash('Nemate dozvolu za pristup ovim podacima', 'error')
        return redirect(url_for('reports.index'))
    
    # Get time entries for the company's projects
    query = db.session.query(TimeEntry).join(Project).filter(Project.company_id == company_id)
    if start_date:
        query = query.filter(TimeEntry.date >= parse_date_from_input(start_date))
    if end_date:
        query = query.filter(TimeEntry.date <= parse_date_from_input(end_date))
    
    entries = query.order_by(TimeEntry.date, Project.name, TimeEntry.user_id).all()
    
    # Get unique dates and projects
    dates = sorted(list(set(entry.date for entry in entries)))
    projects = {}
    for entry in entries:
        if entry.project_id not in projects:
            projects[entry.project_id] = entry.project.name
    
    # Get unique users for the company
    users = {}
    for entry in entries:
        if entry.user_id not in users:
            users[entry.user_id] = entry.user.get_full_name()
    
    # Calculate statistics
    total_hours = sum(entry.hours for entry in entries)
    total_entries = len(entries)
    project_count = len(projects)
    user_count = len(users)
    
    # Calculate earnings
    total_earnings = 0
    user_earnings = {}
    for entry in entries:
        user = User.query.get(entry.user_id)
        hourly_rate = float(user.hourly_rate) if user and user.hourly_rate else 0.0
        entry_earnings = float(entry.hours) * hourly_rate
        total_earnings += entry_earnings
        
        if entry.user_id not in user_earnings:
            user_earnings[entry.user_id] = 0.0
        user_earnings[entry.user_id] += entry_earnings
    
    # Prepare data for template
    company_data = {
        'company': company,
        'total_hours': total_hours,
        'total_entries': total_entries,
        'project_count': project_count,
        'user_count': user_count,
        'total_earnings': total_earnings,
        'dates': dates,
        'projects': projects,
        'users': users,
        'entries': entries,
        'user_earnings': user_earnings,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render_template('reports/company_detail.html', data=company_data)